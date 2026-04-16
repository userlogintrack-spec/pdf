import io
import fitz
from openpyxl import load_workbook


def convert_excel_to_pdf(file_path):
    """
    Convert Excel file to PDF. Each sheet becomes a page.
    """
    wb = load_workbook(file_path, data_only=True)
    pdf_doc = fitz.open()

    page_width = 842  # A4 landscape for tables
    page_height = 595
    margin = 50
    cell_padding = 5
    font_size = 9
    header_font_size = 11
    row_height = font_size * 1.8

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        page = pdf_doc.new_page(width=page_width, height=page_height)
        y_pos = margin

        # Sheet title
        page.insert_text(
            fitz.Point(margin, y_pos + header_font_size),
            sheet_name,
            fontsize=header_font_size,
            fontname="hebo",
        )
        y_pos += header_font_size * 2

        # Calculate column widths
        col_count = ws.max_column or 1
        usable_width = page_width - 2 * margin
        col_width = min(usable_width / col_count, 150)

        # Render rows
        for row in ws.iter_rows(min_row=1, values_only=False):
            if y_pos + row_height > page_height - margin:
                page = pdf_doc.new_page(width=page_width, height=page_height)
                y_pos = margin

            x_pos = margin

            for cell in row:
                value = str(cell.value) if cell.value is not None else ""

                # Truncate long values
                max_chars = int(col_width / (font_size * 0.45))
                if len(value) > max_chars:
                    value = value[:max_chars - 2] + ".."

                # Draw cell border
                rect = fitz.Rect(x_pos, y_pos, x_pos + col_width, y_pos + row_height)
                shape = page.new_shape()
                shape.draw_rect(rect)
                shape.finish(color=(0.7, 0.7, 0.7), width=0.5)
                shape.commit()

                # Draw cell text
                text_y = y_pos + row_height - cell_padding
                page.insert_text(
                    fitz.Point(x_pos + cell_padding, text_y),
                    value,
                    fontsize=font_size,
                    fontname="helv",
                )

                x_pos += col_width

            y_pos += row_height

    output = io.BytesIO()
    pdf_doc.save(output)
    pdf_doc.close()
    wb.close()
    return output.getvalue()
