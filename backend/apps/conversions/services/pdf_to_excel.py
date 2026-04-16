import io
import re
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def convert_pdf_to_excel(file_path):
    """
    Convert PDF to Excel by extracting text and attempting
    to detect tabular data from each page.
    """
    pdf_doc = fitz.open(file_path)
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for page_num in range(pdf_doc.page_count):
        page = pdf_doc[page_num]
        ws = wb.create_sheet(title=f"Page {page_num + 1}")

        # Try to extract tables first
        tables = _extract_tables(page)

        if tables:
            row_offset = 1
            for table in tables:
                for r_idx, row in enumerate(table):
                    for c_idx, cell in enumerate(row):
                        ws.cell(row=row_offset + r_idx, column=c_idx + 1, value=cell)
                row_offset += len(table) + 1  # Gap between tables
        else:
            # Fallback: extract text line by line
            text = page.get_text("text")
            lines = text.strip().split("\n")

            for r_idx, line in enumerate(lines):
                # Try to split by common delimiters
                cells = _split_line(line)
                for c_idx, cell in enumerate(cells):
                    ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell.strip())

        # Auto-fit column widths (approximate)
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    pdf_doc.close()

    return output.getvalue()


def _extract_tables(page):
    """
    Attempt to extract tabular data from a PDF page using
    positional text analysis.
    """
    text_dict = page.get_text("dict")
    blocks = text_dict.get("blocks", [])

    # Collect all text spans with positions
    spans = []
    for block in blocks:
        if block["type"] != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                bbox = span.get("bbox", [0, 0, 0, 0])
                spans.append({
                    "text": span.get("text", "").strip(),
                    "x0": bbox[0],
                    "y0": bbox[1],
                    "x1": bbox[2],
                    "y1": bbox[3],
                })

    if not spans:
        return []

    # Group spans into rows by y-position (within tolerance)
    spans.sort(key=lambda s: (round(s["y0"] / 5) * 5, s["x0"]))

    rows = []
    current_row = [spans[0]]
    current_y = spans[0]["y0"]

    for span in spans[1:]:
        if abs(span["y0"] - current_y) < 5:  # Same row
            current_row.append(span)
        else:
            rows.append(sorted(current_row, key=lambda s: s["x0"]))
            current_row = [span]
            current_y = span["y0"]

    if current_row:
        rows.append(sorted(current_row, key=lambda s: s["x0"]))

    # Check if data looks tabular (multiple rows with similar column count)
    if len(rows) < 2:
        return []

    col_counts = [len(row) for row in rows]
    most_common = max(set(col_counts), key=col_counts.count)

    if most_common < 2:
        return []

    # Extract table data
    table = []
    for row in rows:
        table.append([span["text"] for span in row])

    return [table]


def _split_line(line):
    """Split a line by tabs, multiple spaces, or common delimiters."""
    # Try tab first
    if "\t" in line:
        return line.split("\t")

    # Try multiple spaces (3+)
    parts = re.split(r"\s{3,}", line)
    if len(parts) > 1:
        return parts

    # Try pipe delimiter
    if "|" in line:
        return [p for p in line.split("|") if p.strip()]

    # Return as single cell
    return [line]
